import os
import tkinter as tk
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
#from openpyxl.styles import numbers
from arquivos_xml import processar_xml
from arquivos_csv_unimed import processar_csv_unimed
from arquivos_xlsx_intermedica import processar_xlsx_intermedica
from arquivos_csv_bradesco import processar_csv_bradesco
from tools import centralizar_janela, identificar_convenio,pegar_caminho_onde_abre_o_exe


# função com mensagem de alerta de planilha aberta
def alerta_planilha_aberta():
    janela = tk.Toplevel()
    janela.title("Planilha aberta")

    largura = 350
    altura = 120

    centralizar_janela(largura, altura, janela)
    #janela.geometry(f"{largura}x{altura}")
    janela.resizable(False, False)

    label = tk.Label(
        janela,
        text="A planilha está aberta.\nFeche a planilha e clique em CONTINUAR.",
        pady=20
    )
    label.pack()

    botao = tk.Button(
        janela,
        text="Continuar",
        command=janela.destroy
    )
    botao.pack()

    janela.grab_set()
    janela.wait_window()

def atualizar_planilha(dados, caminho_planilha, nome_aba="FATURADO"):
    dic_erro = {}
    dic_erro[0] = ["OPERADORA", "PROTOCOLO", "GUIA", "MOTIVO"]
    df = pd.read_excel(caminho_planilha, sheet_name=nome_aba, dtype=str) # 24/03 adicionei o dtype

    # Padronizar colunas para string limpa
    df["CONVENIO"] = df["CONVENIO"].astype(str).str.strip()
    df["PROTOCOLO"] = df["PROTOCOLO"].astype(str).str.replace(".0", "", regex=False).str.strip()
    df["Nº GUIA"] = df["Nº GUIA"].astype(str).str.replace(".0", "", regex=False).str.strip()

    if "SENHA" in df.columns:
        df["SENHA"] = df["SENHA"].astype(str).str.replace(".0","",regex=False).str.strip()
    
    linhas_para_atualizar = set()

    for dado in dados:
        convenio = str(dado["convenio"]).strip()
        protocolo = str(dado["protocolo"]).replace(".0", "").strip()
        guia = str(dado.get("guia", "")).replace(".0", "").strip()
        senha = str(dado.get("senha", "")).replace(".0", "").strip() # adicionado para ler o arquivo da intermedica, pois não usa guia e sim a coluna autorização

        # FILTRO DIFERENTE PARA CSV (SEM PROTOCOLO)
        if senha != "":
            filtro = (
                (df["CONVENIO"] == convenio) &
                (df["SENHA"] == senha)
            )

        # FILTRO PADRÃO (OUTROS CONVÊNIOS)
        else:
            if protocolo == "":
                filtro = (
                    (df["CONVENIO"] == convenio) &
                    (df["Nº GUIA"] == guia)
                )
            else:
                filtro = (
                    (df["CONVENIO"] == convenio) &
                    (df["PROTOCOLO"] == protocolo) &
                    (df["Nº GUIA"] == guia)
                )

        if filtro.any():
            linhas_encontradas = df.index[filtro]

                # 🚨 NOVA VALIDAÇÃO AQUI
            serie_valor_pago = df.loc[filtro, "VALOR PAGO"]

            if serie_valor_pago.notna().any() and (serie_valor_pago != 0).any():
                identificador = senha if senha != "" else guia
                dic_erro[f"{convenio}{protocolo}{identificador}"] = [
                    convenio, protocolo, identificador, "VALOR JÁ PREENCHIDO"
                ]
                continue  # 🔥 pula para o próximo registro

            df.loc[filtro, "DATA DE PAGAMENTO"] = str(dado["data_pagamento"])
            df.loc[filtro, "VALOR PAGO"] = dado["valor_pago"]
            df.loc[filtro, "GLOSA"] = dado["glosa"]
            
            vl_pago = dado["valor_pago"] 
            vl_pago = 0 if vl_pago is None else float(vl_pago)
            status = ''
            vl_total = 0 if df.loc[filtro, "VALOR TOTAL"] is None else df.loc[filtro, "VALOR TOTAL"]
            vl_total = float(vl_total.iloc[0])

            if vl_total <= vl_pago:
                status = 'Quitado' 
            elif float(dado["glosa"]) > 0:
                status  = 'Glosa'
            else:                
                status = 'Recebimento parcial'
            
            df.loc[filtro, "STATUS PAGAMENTO"] = status  
        
            # >>> ADICIONADO MOTIVO GLOSA (remove duplicados e evita NAN)
            if "motivo_glosa" in dado and "MOTIVO GLOSA" in df.columns:
                motivo = dado["motivo_glosa"]
                # verificar se está vazio, NaN ou string "nan"
                if pd.isna(motivo) or str(motivo).strip().lower() == "nan" or str(motivo).strip() == "":
                    motivo_final = ""
                else:
                    motivo = str(motivo)
                    # remover motivos repetidos
                    motivos_unicos = list(dict.fromkeys([m.strip() for m in motivo.split("|")]))
                    motivo_final = " | ".join(motivos_unicos)

                df.loc[filtro, "MOTIVO GLOSA"] = str(motivo_final).replace('nan | ','')
            
            # sugestão 
            if len(linhas_encontradas) > 0:
                linhas_para_atualizar.update(linhas_encontradas)

        else:
            identificador = senha if senha != "" else guia
            dic_erro[convenio+protocolo+identificador] = [convenio, protocolo, identificador]

    # salvar mantendo formatação
    while True:
        try:
            wb = load_workbook(caminho_planilha)
            break
        except PermissionError:
            alerta_planilha_aberta()

    ws = wb[nome_aba]
    for i in linhas_para_atualizar:
        for j, col in enumerate(df.columns):
            valor = df.iloc[i, j]
            ws.cell(row=i+2, column=j+1).value = valor

    # REAPLICAR FORMATO DE MOEDA
    colunas_moeda = []
    for idx, col in enumerate(df.columns):
        if col in ["VALOR PAGO", "GLOSA"]:
            colunas_moeda.append(idx + 1)

    for col in colunas_moeda:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = 'R$ #,##0.00'

    while True:
        try:
            wb.save(caminho_planilha)
            break
        except PermissionError:
            alerta_planilha_aberta()

    return dic_erro

def pegar_data_hora_atual():
    agora = datetime.now()
    return agora.strftime("%d%m%Y_%H%M%S")

def gravar_log_txt(dic_log, arquivo):
    log = f"log_{pegar_data_hora_atual()}.log"
    arquivo = os.path.join(arquivo, log)
    with open(arquivo, "a", encoding="utf-8") as f:
        for _, valor in dic_log.items():
            f.write(f"{valor}\n")
        f.write("-" * 40 + "\n")

def gravar_log_xlsx(dic_log, arquivo):
    log = f"log_{pegar_data_hora_atual()}.xlsx"
    arquivo = os.path.join(arquivo, log)
    #df = pd.DataFrame(dic_log)  # transforma em uma linha
    df = pd.DataFrame.from_dict(dic_log, orient="index")
    df.to_excel(arquivo, sheet_name="log", index=False)


def executar_leitura_arquivos(caminho_planilha, pasta_raiz_arquivos=r"arquivos", salvar_log_erro=r"."):
    try:
        # Pastas de entrada
        pasta_logs = os.path.join(salvar_log_erro, "logs")

        PASTA_XML = os.path.join(pasta_raiz_arquivos, "arquivos/XML")
        PASTA_CSV = os.path.join(pasta_raiz_arquivos, "arquivos/CSV")
        PASTA_EXCEL = os.path.join(pasta_raiz_arquivos, "arquivos/EXCEL")
        
        # Lista de todas as pastas de entrada
        PASTAS_ENTRADA = [PASTA_XML, PASTA_CSV, PASTA_EXCEL, pasta_logs]

        # Criar pastas automaticamente se não existirem
        for pasta in PASTAS_ENTRADA:
            os.makedirs(pasta, exist_ok=True)

        # Nome da aba
        nome_aba_planilha = "FATURADO"

        # adicionado para ler o arquivo csv da unimed, caso seja identificado o convenio unimed no nome do arquivo 
        dic_processadores_csv = {
            "UNIMED": processar_csv_unimed
            ,"BRADESCO": processar_csv_bradesco
        }

        dic_erro = {}
        lst_dados_palnilha = []
        dados = []

        for pasta in PASTAS_ENTRADA:
            for arquivo in os.listdir(pasta):
                caminho = os.path.join(pasta, arquivo)
                extencao_arquivo = Path(caminho).suffix.upper()
                dados = []

                if os.path.isfile(caminho):
                    # dados = processar_arquivo(caminho, dic_processador_extensao)
                    if extencao_arquivo == '.XML':
                        dados = processar_xml(caminho)
                    
                    if extencao_arquivo == '.CSV':
                        convenio = identificar_convenio(caminho)

                        if convenio in dic_processadores_csv:
                            dados = dic_processadores_csv[convenio](caminho)
                    
                    if extencao_arquivo == '.XLSX':
                        dados = processar_xlsx_intermedica(caminho)

                    if dados is not None:
                        lst_dados_palnilha += dados

        dic_erro = atualizar_planilha(lst_dados_palnilha, caminho_planilha, nome_aba_planilha)
        gravar_log_xlsx(dic_erro, pasta_logs)

        del dic_erro
        del dados
        del lst_dados_palnilha

        return 1
    except Exception as error:
        return error